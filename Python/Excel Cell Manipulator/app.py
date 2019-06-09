"""
To be used when using this module as a separate
method in  larger application development

-> Change the parameter wb = excel,load_workbook(filename)
-> make the same change where the final file is saved
-------------------------------------
def process_excel_workbook(filename):
-------------------------------------
"""

import openpyxl as excel
from openpyxl.chart import BarChart, Reference

wb = excel.load_workbook("transactions.xlsx")
sheet = wb['Sheet1']

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

values = Reference(sheet,
          min_row = 2,
          max_row = sheet.max_row,
          min_col = 4,
          max_col = 4)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'E2')

wb.save('Updated_tranasactions.xlsx')